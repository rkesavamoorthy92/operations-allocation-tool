from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.domain.exceptions import InvalidQcResultError, InvalidQcRuleError, InvalidStateTransitionError
from operations_allocation.domain.models import RunState
from operations_allocation.infrastructure.file_artifacts import FileArtifactManager
from operations_allocation.infrastructure.xlsx_writer import write_associate_workbook
from operations_allocation.persistence.database import Database
from operations_allocation.persistence.repositories import (
    AllocationResultRepository,
    ArtifactRepository,
    AuditRepository,
    EligiblePopulationRepository,
    ManifestRepository,
    ProgramRepository,
    RunRepository,
    SamplingResultRepository,
    SnapshotRepository,
)
from operations_allocation.services.allocation import AllocationService
from operations_allocation.services.audit import AuditService
from operations_allocation.services.consolidation import ConsolidationService
from operations_allocation.services.distribution import DistributionService
from operations_allocation.services.eligible_population import EligiblePopulationService
from operations_allocation.services.program_configuration import ProgramConfigurationService
from operations_allocation.services.qc import QcService
from operations_allocation.services.run_orchestration import RunOrchestrationService
from operations_allocation.services.sampling import SamplingService
from operations_allocation.services.source_import import SourceImportService
from tests.unit.test_distribution_service import distribution_config


def qc_config() -> dict:
    config = distribution_config()
    config["qc"] = {
        "rules": [
            {"name": "qc_score", "rule_type": "ratio_percentage", "numerator": "pass_count", "denominator": "audited_count"},
            {"name": "error_rate", "rule_type": "ratio_percentage", "numerator": "fail_count", "denominator": "audited_count"},
        ]
    }
    return config


class QcServiceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        database = Database(Path(self.tempdir.name) / "qc.db")
        database.initialize_schema()
        programs = ProgramRepository(database)
        self.runs = RunRepository(database)
        self.snapshots = SnapshotRepository(database)
        self.audit_repository = AuditRepository(database)
        audit = AuditService(self.audit_repository, "Test Application")
        ProgramConfigurationService(programs).create_program("MX-PT", "MX PT")
        ProgramConfigurationService(programs).save_version(qc_config())

        self.orchestration = RunOrchestrationService(runs=self.runs, snapshots=self.snapshots, manifests=ManifestRepository(database), audit=audit)
        self.file_artifacts = FileArtifactManager(base_directory=Path(self.tempdir.name) / "artifacts", artifacts=ArtifactRepository(database))
        self.source_import = SourceImportService(snapshots=self.snapshots, file_artifacts=self.file_artifacts, audit=audit)
        populations = EligiblePopulationRepository(database)
        self.eligible_population_service = EligiblePopulationService(runs=self.runs, snapshots=self.snapshots, populations=populations, audit=audit)
        self.sampling_results = SamplingResultRepository(database)
        self.sampling_service = SamplingService(runs=self.runs, snapshots=self.snapshots, populations=populations, sampling_results=self.sampling_results, audit=audit)
        self.allocation_results = AllocationResultRepository(database)
        self.allocation_service = AllocationService(runs=self.runs, snapshots=self.snapshots, sampling_results=self.sampling_results, allocation_results=self.allocation_results, audit=audit)
        self.distribution_service = DistributionService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, source_import=self.source_import, file_artifacts=self.file_artifacts, audit=audit)
        self.consolidation_service = ConsolidationService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, file_artifacts=self.file_artifacts, audit=audit)
        self.qc_service = QcService(runs=self.runs, snapshots=self.snapshots, file_artifacts=self.file_artifacts, audit=audit)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_source_csv(self, row_count: int) -> Path:
        path = Path(self.tempdir.name) / "source.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, row_count + 1):
                writer.writerow([f"P{i:03d}", "Shoes"])
        return path

    def _run_to_consolidated(self, *, associates: list[dict], sample_count: int, row_count: int, configuration: dict | None = None) -> str:
        config = configuration if configuration is not None else qc_config()
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=config, sampling={"method": "count", "value": sample_count}, random_seed="seed", associates=associates)
        source_path = self._write_source_csv(row_count)
        canonical_rows, _ = self.source_import.import_source(run_id=run.run_id, file_path=source_path)
        self.eligible_population_service.validate(run_id=run.run_id, rows=canonical_rows)
        self.eligible_population_service.freeze(run_id=run.run_id, rows=canonical_rows)
        self.sampling_service.sample(run_id=run.run_id)
        self.allocation_service.finalize(run_id=run.run_id)
        artifacts = self.distribution_service.distribute(run_id=run.run_id)

        returned_dir = Path(self.tempdir.name) / f"returned_{run.run_id}"
        returned_dir.mkdir()
        returned_paths = []
        for artifact in artifacts:
            path = self.file_artifacts.run_directory(run.run_id) / artifact.relative_path
            workbook = load_workbook(path, read_only=True, data_only=True)
            metadata = {row[0]: row[1] for row in workbook["Metadata"].iter_rows(min_row=2, values_only=True) if row[0] is not None}
            data_iter = workbook["Data"].iter_rows(values_only=True)
            headers = next(data_iter)
            rows = [list(row) for row in data_iter]
            workbook.close()
            returned_path = returned_dir / path.name
            returned_path.write_bytes(write_associate_workbook(metadata=metadata, headers=headers, rows=rows))
            returned_paths.append((returned_path, artifact.associate_id))

        self.consolidation_service.import_returned_files(run_id=run.run_id, files=returned_paths)
        self.consolidation_service.finalize(run_id=run.run_id)
        return run.run_id

    def _write_qc_report(self, rows: list[tuple[str, str, str]]) -> Path:
        path = Path(self.tempdir.name) / "qc_report.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "Allocated To", "QC Result"])
            for row in rows:
                writer.writerow(list(row))
        return path

    def test_spec_worked_example_end_to_end(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10}]
        run_id = self._run_to_consolidated(associates=associates, sample_count=10, row_count=10)
        qc_rows = [(f"P{i:03d}", "A001", "Pass") for i in range(1, 9)] + [(f"P{i:03d}", "A001", "Fail") for i in range(9, 11)]
        qc_report_path = self._write_qc_report(qc_rows)

        report = self.qc_service.import_and_evaluate(run_id=run_id, file_path=qc_report_path)
        self.assertEqual(report.run_metrics["qc_score"].value, Decimal(80))
        self.assertEqual(report.run_metrics["error_rate"].value, Decimal(20))
        self.assertEqual(self.runs.get(run_id).state, RunState.QC_COMPLETED)

    def test_per_associate_breakdown(self) -> None:
        associates = [
            {"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5},
            {"associate_id": "A002", "name": "John Smith", "email": "john.smith@example.test", "active": True, "target": 5, "maximum_capacity": 5},
        ]
        run_id = self._run_to_consolidated(associates=associates, sample_count=10, row_count=10)
        allocation = self.allocation_results.get(run_id)
        rows = []
        for assignment in allocation.assignments:
            outcome = "Pass" if assignment.associate_id == "A001" else "Fail"
            for identifier in assignment.assigned_identifiers:
                rows.append((identifier, assignment.associate_id, outcome))
        qc_report_path = self._write_qc_report(rows)

        report = self.qc_service.import_and_evaluate(run_id=run_id, file_path=qc_report_path)
        self.assertEqual(report.associate_metrics["A001"]["qc_score"].value, Decimal(100))
        self.assertEqual(report.associate_metrics["A002"]["qc_score"].value, Decimal(0))

    def test_zero_denominator_is_not_applicable(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5}]
        run_id = self._run_to_consolidated(associates=associates, sample_count=5, row_count=5)
        qc_report_path = self._write_qc_report([])
        report = self.qc_service.import_and_evaluate(run_id=run_id, file_path=qc_report_path)
        self.assertTrue(report.run_metrics["qc_score"].is_not_applicable)

    def test_missing_qc_rules_raises_clear_error(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5}]
        run_id = self._run_to_consolidated(associates=associates, sample_count=5, row_count=5, configuration=distribution_config())
        qc_report_path = self._write_qc_report([("P001", "A001", "Pass")])
        with self.assertRaises(InvalidQcRuleError):
            self.qc_service.import_and_evaluate(run_id=run_id, file_path=qc_report_path)

    def test_unrecognized_outcome_value_raises(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5}]
        run_id = self._run_to_consolidated(associates=associates, sample_count=5, row_count=5)
        qc_report_path = self._write_qc_report([("P001", "A001", "Maybe")])
        with self.assertRaises(InvalidQcResultError):
            self.qc_service.import_and_evaluate(run_id=run_id, file_path=qc_report_path)

    def test_requires_consolidated_state(self) -> None:
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=qc_config(), sampling={"method": "count", "value": 1}, random_seed="seed", associates=[])
        qc_report_path = self._write_qc_report([])
        with self.assertRaises(InvalidStateTransitionError):
            self.qc_service.import_and_evaluate(run_id=run.run_id, file_path=qc_report_path)

    def test_records_audit_event(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5}]
        run_id = self._run_to_consolidated(associates=associates, sample_count=5, row_count=5)
        qc_report_path = self._write_qc_report([("P001", "A001", "Pass")])
        self.qc_service.import_and_evaluate(run_id=run_id, file_path=qc_report_path)
        actions = [event["action"] for event in self.audit_repository.for_run(run_id)]
        self.assertIn("RUN_QC_COMPLETED", actions)
