from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.infrastructure.file_artifacts import FileArtifactManager
from operations_allocation.persistence.database import Database
from operations_allocation.persistence.repositories import (
    AllocationResultRepository,
    ArtifactRepository,
    AuditRepository,
    EligiblePopulationRepository,
    ManifestRepository,
    ProgramRepository,
    RunRepository,
    SamplingResultRepository,
    SnapshotRepository,
)
from operations_allocation.services.allocation import AllocationService
from operations_allocation.services.audit import AuditService
from operations_allocation.services.eligible_population import EligiblePopulationService
from operations_allocation.services.insights import InsightsService
from operations_allocation.services.program_configuration import ProgramConfigurationService
from operations_allocation.services.reporting import ReportingService
from operations_allocation.services.run_orchestration import RunOrchestrationService
from operations_allocation.services.sampling import SamplingService
from operations_allocation.services.source_import import SourceImportService
from tests.unit.test_distribution_service import distribution_config


class ReportingServiceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        database = Database(Path(self.tempdir.name) / "reporting.db")
        database.initialize_schema()
        self.programs = ProgramRepository(database)
        self.runs = RunRepository(database)
        snapshots = SnapshotRepository(database)
        audit = AuditService(AuditRepository(database), "Test Application")
        ProgramConfigurationService(self.programs).create_program("MX-PT", "MX PT")
        ProgramConfigurationService(self.programs).save_version(distribution_config())

        orchestration = RunOrchestrationService(runs=self.runs, snapshots=snapshots, manifests=ManifestRepository(database), audit=audit)
        file_artifacts = FileArtifactManager(base_directory=Path(self.tempdir.name) / "artifacts", artifacts=ArtifactRepository(database))
        source_import = SourceImportService(snapshots=snapshots, file_artifacts=file_artifacts, audit=audit)
        populations = EligiblePopulationRepository(database)
        eligible_population = EligiblePopulationService(runs=self.runs, snapshots=snapshots, populations=populations, audit=audit)
        sampling_results = SamplingResultRepository(database)
        sampling = SamplingService(runs=self.runs, snapshots=snapshots, populations=populations, sampling_results=sampling_results, audit=audit)
        allocation_results = AllocationResultRepository(database)
        allocation = AllocationService(runs=self.runs, snapshots=snapshots, sampling_results=sampling_results, allocation_results=allocation_results, audit=audit)
        insights = InsightsService(runs=self.runs, allocation_results=allocation_results, file_artifacts=file_artifacts)
        self.reporting = ReportingService(runs=self.runs, programs=self.programs, insights=insights)

        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10}]
        self.run = orchestration.create_run(program_id="MX-PT", created_by="tester", created_on=date(2026, 8, 15))
        orchestration.freeze_setup(run_id=self.run.run_id, program_configuration=distribution_config(), sampling={"method": "count", "value": 10}, random_seed="seed", associates=associates)

        source_path = Path(self.tempdir.name) / "source.csv"
        with source_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, 11):
                writer.writerow([f"P{i:03d}", "Shoes"])
        canonical_rows, _ = source_import.import_source(run_id=self.run.run_id, file_path=source_path)
        eligible_population.validate(run_id=self.run.run_id, rows=canonical_rows)
        eligible_population.freeze(run_id=self.run.run_id, rows=canonical_rows)
        sampling.sample(run_id=self.run.run_id)
        allocation.finalize(run_id=self.run.run_id)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def test_export_run_summary_produces_workbook_with_all_sheets(self) -> None:
        content = self.reporting.export_run_summary(run_id=self.run.run_id)
        export_path = Path(self.tempdir.name) / "summary.xlsx"
        export_path.write_bytes(content)
        workbook = load_workbook(export_path, read_only=True, data_only=True)
        self.assertEqual(workbook.sheetnames, ["Summary", "Allocation Utilization", "Associate QC Performance", "Top Error Categories"])
        workbook.close()

    def test_summary_sheet_reflects_run_and_program_identity(self) -> None:
        content = self.reporting.export_run_summary(run_id=self.run.run_id)
        export_path = Path(self.tempdir.name) / "summary2.xlsx"
        export_path.write_bytes(content)
        workbook = load_workbook(export_path, read_only=True, data_only=True)
        rows = {row[0]: row[1] for row in workbook["Summary"].iter_rows(values_only=True)}
        workbook.close()
        self.assertEqual(rows["Run ID"], self.run.run_id)
        self.assertEqual(rows["Program"], "MX PT (MX-PT)")
        self.assertEqual(rows["Has Prior Completed Run for Comparison"], False)

    def test_utilization_sheet_reflects_the_finalized_allocation(self) -> None:
        content = self.reporting.export_run_summary(run_id=self.run.run_id)
        export_path = Path(self.tempdir.name) / "summary3.xlsx"
        export_path.write_bytes(content)
        workbook = load_workbook(export_path, read_only=True, data_only=True)
        rows = dict(workbook["Allocation Utilization"].iter_rows(min_row=2, values_only=True))
        workbook.close()
        self.assertEqual(rows["A001"], 100.0)
