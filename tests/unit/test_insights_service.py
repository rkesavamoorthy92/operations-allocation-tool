from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.domain.models import RunState
from operations_allocation.infrastructure.xlsx_writer import write_associate_workbook
from operations_allocation.ui import run_actions
from operations_allocation.ui.app_context import AppContext
from tests.unit.test_qc_service import qc_config


class InsightsServiceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.context = AppContext.build(data_directory=Path(self.tempdir.name) / "app_data")
        self.context.program_configuration.create_program("MX-PT", "MX PT")
        self.context.program_configuration.save_version(qc_config())
        self.insights = self.context.insights

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_source_csv(self, name: str, row_count: int) -> Path:
        path = Path(self.tempdir.name) / name
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, row_count + 1):
                writer.writerow([f"P{i:03d}", "Shoes"])
        return path

    def _run_to_completed(self, *, source_name: str, pass_count: int, fail_count: int) -> str:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10}]
        run = self.context.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.context.orchestration.freeze_setup(run_id=run.run_id, program_configuration=qc_config(), sampling={"method": "count", "value": 10}, random_seed="seed", associates=associates)
        canonical_rows, _ = run_actions.import_source_and_validate(self.context, run_id=run.run_id, file_path=self._write_source_csv(source_name, 10))
        run_actions.freeze_eligible_population(self.context, run_id=run.run_id, canonical_rows=canonical_rows)
        run_actions.sample(self.context, run_id=run.run_id)
        run_actions.finalize_allocation(self.context, run_id=run.run_id)
        artifacts = run_actions.distribute(self.context, run_id=run.run_id)

        path = self.context.file_artifacts.run_directory(run.run_id) / artifacts[0].relative_path
        workbook = load_workbook(path, read_only=True, data_only=True)
        metadata = {r[0]: r[1] for r in workbook["Metadata"].iter_rows(min_row=2, values_only=True) if r[0] is not None}
        data_iter = workbook["Data"].iter_rows(values_only=True)
        headers = next(data_iter)
        rows = [list(r) for r in data_iter]
        workbook.close()
        returned_dir = Path(self.tempdir.name) / f"returned_{source_name}"
        returned_dir.mkdir()
        returned_path = returned_dir / path.name
        returned_path.write_bytes(write_associate_workbook(metadata=metadata, headers=headers, rows=rows))

        run_actions.import_returned_files(self.context, run_id=run.run_id, files=[(returned_path, "A001")])
        run_actions.finalize_consolidation(self.context, run_id=run.run_id)

        qc_report_path = Path(self.tempdir.name) / f"{source_name}_qc.csv"
        with qc_report_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "Allocated To", "QC Result"])
            for i, row in enumerate(rows):
                outcome = "Pass" if i < pass_count else "Fail"
                writer.writerow([row[0], row[3], outcome])
        run_actions.import_qc_report(self.context, run_id=run.run_id, file_path=qc_report_path)
        run_actions.generate_errors(self.context, run_id=run.run_id)
        run_actions.complete_run(self.context, run_id=run.run_id)
        return run.run_id

    def test_no_previous_completed_run_is_not_applicable(self) -> None:
        run_id = self._run_to_completed(source_name="first.csv", pass_count=8, fail_count=2)
        report = self.insights.generate(run_id=run_id)
        self.assertFalse(report.historical.is_applicable)
        self.assertEqual(report.historical.current_qc_score, Decimal("80.00"))

    def test_second_run_compares_against_first(self) -> None:
        self._run_to_completed(source_name="first.csv", pass_count=9, fail_count=1)
        second_run_id = self._run_to_completed(source_name="second.csv", pass_count=8, fail_count=2)
        report = self.insights.generate(run_id=second_run_id)
        self.assertTrue(report.historical.is_applicable)
        self.assertEqual(report.historical.previous_qc_score, Decimal("90.00"))
        self.assertEqual(report.historical.current_qc_score, Decimal("80.00"))
        self.assertEqual(report.historical.qc_score_change_points, Decimal("-10.00"))

    def test_allocation_utilization_and_completion_rate(self) -> None:
        run_id = self._run_to_completed(source_name="first.csv", pass_count=10, fail_count=0)
        report = self.insights.generate(run_id=run_id)
        self.assertEqual(report.allocation_utilization["A001"], Decimal("100.00"))
        self.assertEqual(report.completion_rate, Decimal("100.00"))

    def test_error_categories_present_after_generation(self) -> None:
        run_id = self._run_to_completed(source_name="first.csv", pass_count=8, fail_count=2)
        report = self.insights.generate(run_id=run_id)
        self.assertIsInstance(report.top_error_categories, tuple)

    def test_insights_for_run_missing_optional_stages_returns_gracefully(self) -> None:
        run = self.context.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        report = self.insights.generate(run_id=run.run_id)
        self.assertEqual(report.allocation_utilization, {})
        self.assertIsNone(report.completion_rate)
        self.assertFalse(report.historical.is_applicable)
