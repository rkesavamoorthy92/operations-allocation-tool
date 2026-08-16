from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.domain.exceptions import InvalidErrorRecordError
from operations_allocation.domain.models import ErrorSource
from operations_allocation.infrastructure.file_artifacts import FileArtifactManager
from operations_allocation.infrastructure.xlsx_writer import write_associate_workbook
from operations_allocation.persistence.database import Database
from operations_allocation.persistence.repositories import (
    AllocationResultRepository,
    ArtifactRepository,
    AuditRepository,
    EligiblePopulationRepository,
    ManifestRepository,
    ProgramRepository,
    RunRepository,
    SamplingResultRepository,
    SnapshotRepository,
)
from operations_allocation.services.allocation import AllocationService
from operations_allocation.services.audit import AuditService
from operations_allocation.services.consolidation import ConsolidationService
from operations_allocation.services.distribution import DistributionService
from operations_allocation.services.eligible_population import EligiblePopulationService
from operations_allocation.services.errors import ErrorService
from operations_allocation.services.program_configuration import ProgramConfigurationService
from operations_allocation.services.run_orchestration import RunOrchestrationService
from operations_allocation.services.sampling import SamplingService
from operations_allocation.services.source_import import SourceImportService
from tests.unit.test_distribution_service import distribution_config


def errors_config() -> dict:
    config = distribution_config()
    config["errors"] = {
        "categories": ["Missing", "Duplicate", "Unexpected"],
        "classification_rules": [
            {"match": {"disposition": "missing"}, "category": "Missing", "type": "Item Not Returned", "severity": "Critical"},
            {"match": {"disposition": "duplicate"}, "category": "Duplicate", "type": "Duplicate Return", "severity": "High"},
            {"match": {"disposition": "unexpected"}, "category": "Unexpected", "type": "Unallocated Item Returned", "severity": "Medium"},
            {"match": {"disposition": "wrong_associate"}, "category": "Misrouted", "type": "Wrong Associate", "severity": "High"},
        ],
    }
    return config


class ErrorServiceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        database = Database(Path(self.tempdir.name) / "errors.db")
        database.initialize_schema()
        programs = ProgramRepository(database)
        self.runs = RunRepository(database)
        self.snapshots = SnapshotRepository(database)
        self.audit_repository = AuditRepository(database)
        audit = AuditService(self.audit_repository, "Test Application")
        ProgramConfigurationService(programs).create_program("MX-PT", "MX PT")
        ProgramConfigurationService(programs).save_version(errors_config())

        self.orchestration = RunOrchestrationService(runs=self.runs, snapshots=self.snapshots, manifests=ManifestRepository(database), audit=audit)
        self.file_artifacts = FileArtifactManager(base_directory=Path(self.tempdir.name) / "artifacts", artifacts=ArtifactRepository(database))
        self.source_import = SourceImportService(snapshots=self.snapshots, file_artifacts=self.file_artifacts, audit=audit)
        populations = EligiblePopulationRepository(database)
        self.eligible_population_service = EligiblePopulationService(runs=self.runs, snapshots=self.snapshots, populations=populations, audit=audit)
        self.sampling_results = SamplingResultRepository(database)
        self.sampling_service = SamplingService(runs=self.runs, snapshots=self.snapshots, populations=populations, sampling_results=self.sampling_results, audit=audit)
        self.allocation_results = AllocationResultRepository(database)
        self.allocation_service = AllocationService(runs=self.runs, snapshots=self.snapshots, sampling_results=self.sampling_results, allocation_results=self.allocation_results, audit=audit)
        self.distribution_service = DistributionService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, source_import=self.source_import, file_artifacts=self.file_artifacts, audit=audit)
        self.consolidation_service = ConsolidationService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, file_artifacts=self.file_artifacts, audit=audit)
        self.error_service = ErrorService(snapshots=self.snapshots, file_artifacts=self.file_artifacts, audit=audit)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_source_csv(self, row_count: int) -> Path:
        path = Path(self.tempdir.name) / "source.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, row_count + 1):
                writer.writerow([f"P{i:03d}", "Shoes"])
        return path

    def _run_with_exceptions(self) -> str:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10}]
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=errors_config(), sampling={"method": "count", "value": 10}, random_seed="seed", associates=associates)
        source_path = self._write_source_csv(10)
        canonical_rows, _ = self.source_import.import_source(run_id=run.run_id, file_path=source_path)
        self.eligible_population_service.validate(run_id=run.run_id, rows=canonical_rows)
        self.eligible_population_service.freeze(run_id=run.run_id, rows=canonical_rows)
        self.sampling_service.sample(run_id=run.run_id)
        self.allocation_service.finalize(run_id=run.run_id)
        artifacts = self.distribution_service.distribute(run_id=run.run_id)
        path = self.file_artifacts.run_directory(run.run_id) / artifacts[0].relative_path

        workbook = load_workbook(path, read_only=True, data_only=True)
        metadata = {row[0]: row[1] for row in workbook["Metadata"].iter_rows(min_row=2, values_only=True) if row[0] is not None}
        data_iter = workbook["Data"].iter_rows(values_only=True)
        headers = next(data_iter)
        rows = [list(row) for row in data_iter]
        workbook.close()

        rows = rows[:-1]  # Drop one -> missing.
        rows.append(list(rows[0]))  # Duplicate.
        rows.append(["ZZZ-UNKNOWN", "Shoes", None, "A001", run.run_id])  # Unexpected.

        returned_dir = Path(self.tempdir.name) / "returned"
        returned_dir.mkdir()
        returned_path = returned_dir / path.name
        returned_path.write_bytes(write_associate_workbook(metadata=metadata, headers=headers, rows=rows))

        self.consolidation_service.import_returned_files(run_id=run.run_id, files=[(returned_path, "A001")])
        self.consolidation_service.finalize(run_id=run.run_id, override=True, overridden_by="qa_lead", override_reason="Testing error generation.")
        return run.run_id

    def test_generate_from_consolidation_classifies_each_exception_type(self) -> None:
        run_id = self._run_with_exceptions()
        records = self.error_service.generate_from_consolidation(run_id=run_id)
        categories = {r.category for r in records}
        self.assertIn("Missing", categories)
        self.assertIn("Duplicate", categories)
        self.assertIn("Unexpected", categories)
        self.assertTrue(all(r.source == ErrorSource.GENERATED for r in records))

    def test_generate_records_audit_event(self) -> None:
        run_id = self._run_with_exceptions()
        self.error_service.generate_from_consolidation(run_id=run_id)
        actions = [event["action"] for event in self.audit_repository.for_run(run_id)]
        self.assertIn("ERRORS_GENERATED", actions)

    def test_import_errors_trusts_pre_classified_rows(self) -> None:
        run_id = self._run_with_exceptions()
        error_report = Path(self.tempdir.name) / "error_report.csv"
        with error_report.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "Allocated To", "Category", "Type", "Severity"])
            writer.writerow(["P001", "A001", "Custom Category", "Custom Type", "Low"])
        records = self.error_service.import_errors(run_id=run_id, file_path=error_report)
        self.assertEqual(records[0].category, "Custom Category")
        self.assertEqual(records[0].source, ErrorSource.IMPORTED)

    def test_import_errors_missing_identifier_raises(self) -> None:
        run_id = self._run_with_exceptions()
        error_report = Path(self.tempdir.name) / "bad_report.csv"
        with error_report.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "Allocated To"])
            writer.writerow(["", "A001"])
        with self.assertRaises(InvalidErrorRecordError):
            self.error_service.import_errors(run_id=run_id, file_path=error_report)

    def test_list_records_is_empty_before_any_errors_generated_or_imported(self) -> None:
        run_id = self._run_with_exceptions()
        self.assertEqual(self.error_service.list_records(run_id=run_id), ())

    def test_list_records_returns_both_generated_and_imported(self) -> None:
        run_id = self._run_with_exceptions()
        generated = self.error_service.generate_from_consolidation(run_id=run_id)

        error_report = Path(self.tempdir.name) / "error_report.csv"
        with error_report.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "Allocated To", "Category", "Type", "Severity"])
            writer.writerow(["P001", "A001", "Custom Category", "Custom Type", "Low"])
        imported = self.error_service.import_errors(run_id=run_id, file_path=error_report)

        records = self.error_service.list_records(run_id=run_id)
        self.assertEqual(len(records), len(generated) + len(imported))
        self.assertEqual({r.source for r in records}, {ErrorSource.GENERATED, ErrorSource.IMPORTED})

    def test_export_report_produces_a_readable_workbook_with_every_record(self) -> None:
        run_id = self._run_with_exceptions()
        self.error_service.generate_from_consolidation(run_id=run_id)

        content = self.error_service.export_report(run_id=run_id)

        export_path = Path(self.tempdir.name) / "error_report_export.xlsx"
        export_path.write_bytes(content)
        workbook = load_workbook(export_path, read_only=True, data_only=True)
        self.assertIn("Errors", workbook.sheetnames)
        sheet = workbook["Errors"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        header_row = rows[0]
        self.assertEqual(header_row[:6], ("Identifier", "Associate ID", "Category", "Type", "Severity", "Source"))
        self.assertEqual(len(rows) - 1, len(self.error_service.list_records(run_id=run_id)))

    def test_export_report_with_no_errors_yet_is_an_empty_but_valid_workbook(self) -> None:
        run_id = self._run_with_exceptions()
        content = self.error_service.export_report(run_id=run_id)
        export_path = Path(self.tempdir.name) / "empty_export.xlsx"
        export_path.write_bytes(content)
        workbook = load_workbook(export_path, read_only=True, data_only=True)
        sheet = workbook["Errors"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        self.assertEqual(len(rows), 1)  # Header row only.
