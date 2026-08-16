from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.domain.models import RunState
from operations_allocation.infrastructure.xlsx_writer import write_associate_workbook
from operations_allocation.ui import run_actions
from operations_allocation.ui.app_context import AppContext
from tests.unit.test_qc_service import qc_config


class RunActionsTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.context = AppContext.build(data_directory=Path(self.tempdir.name) / "app_data")
        self.context.program_configuration.create_program("MX-PT", "MX PT")
        self.context.program_configuration.save_version(qc_config())

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_source_csv(self, row_count: int) -> Path:
        path = Path(self.tempdir.name) / "source.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, row_count + 1):
                writer.writerow([f"P{i:03d}", "Shoes"])
        return path

    def test_full_lifecycle_via_the_facade(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10}]
        run = self.context.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.context.orchestration.freeze_setup(run_id=run.run_id, program_configuration=qc_config(), sampling={"method": "count", "value": 10}, random_seed="seed", associates=associates)

        canonical_rows, summary = run_actions.import_source_and_validate(self.context, run_id=run.run_id, file_path=self._write_source_csv(10))
        self.assertEqual(summary.total_rows, 10)
        run_actions.freeze_eligible_population(self.context, run_id=run.run_id, canonical_rows=canonical_rows)
        run_actions.sample(self.context, run_id=run.run_id)

        plan = run_actions.preview_allocation(self.context, run_id=run.run_id)
        self.assertIsNotNone(plan)
        result = run_actions.finalize_allocation(self.context, run_id=run.run_id)
        self.assertEqual(len(result.assignments), 1)

        artifacts = run_actions.distribute(self.context, run_id=run.run_id)
        self.assertEqual(len(artifacts), 1)

        # Round-trip a perfect return through the facade.
        path = self.context.file_artifacts.run_directory(run.run_id) / artifacts[0].relative_path
        workbook = load_workbook(path, read_only=True, data_only=True)
        metadata = {r[0]: r[1] for r in workbook["Metadata"].iter_rows(min_row=2, values_only=True) if r[0] is not None}
        data_iter = workbook["Data"].iter_rows(values_only=True)
        headers = next(data_iter)
        rows = [list(r) for r in data_iter]
        workbook.close()
        returned_path = Path(self.tempdir.name) / path.name
        returned_path.write_bytes(write_associate_workbook(metadata=metadata, headers=headers, rows=rows))

        run_actions.import_returned_files(self.context, run_id=run.run_id, files=[(returned_path, "A001")])
        run_actions.finalize_consolidation(self.context, run_id=run.run_id)
        self.assertEqual(self.context.runs.get(run.run_id).state, RunState.CONSOLIDATED)

        qc_report_path = Path(self.tempdir.name) / "qc_report.csv"
        with qc_report_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "Allocated To", "QC Result"])
            for row in rows:
                writer.writerow([row[0], row[3], "Pass"])
        report = run_actions.import_qc_report(self.context, run_id=run.run_id, file_path=qc_report_path)
        self.assertFalse(report.run_metrics["qc_score"].is_not_applicable)

        run_actions.generate_errors(self.context, run_id=run.run_id)

        export_content = run_actions.export_error_report(self.context, run_id=run.run_id)
        export_path = Path(self.tempdir.name) / "error_report.xlsx"
        export_path.write_bytes(export_content)
        export_workbook = load_workbook(export_path, read_only=True, data_only=True)
        self.assertIn("Errors", export_workbook.sheetnames)
        export_workbook.close()

        completed = run_actions.complete_run(self.context, run_id=run.run_id)
        self.assertEqual(completed.state, RunState.COMPLETED)
