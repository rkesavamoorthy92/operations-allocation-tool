from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.domain.exceptions import InvalidStateTransitionError
from operations_allocation.domain.models import ArtifactType, RunState
from operations_allocation.infrastructure.file_artifacts import FileArtifactManager
from operations_allocation.persistence.database import Database
from operations_allocation.persistence.repositories import (
    AllocationResultRepository,
    ArtifactRepository,
    AuditRepository,
    EligiblePopulationRepository,
    ManifestRepository,
    ProgramRepository,
    RunRepository,
    SamplingResultRepository,
    SnapshotRepository,
)
from operations_allocation.services.allocation import AllocationService
from operations_allocation.services.audit import AuditService
from operations_allocation.services.distribution import DistributionService
from operations_allocation.services.eligible_population import EligiblePopulationService
from operations_allocation.services.program_configuration import ProgramConfigurationService
from operations_allocation.services.run_orchestration import RunOrchestrationService
from operations_allocation.services.sampling import SamplingService
from operations_allocation.services.source_import import SourceImportService
from tests.unit.test_foundation import valid_config


def distribution_config() -> dict:
    config = valid_config()
    config["fields"] = [
        {"name": "product_id", "source_column": "Product ID", "ownership": "source", "data_type": "string", "required": True, "output_order": 0},
        {"name": "pt", "source_column": "PT", "ownership": "source", "data_type": "string", "required": False, "output_order": 1},
        {"name": "partner_feedback", "ownership": "response", "data_type": "string", "required": False, "output_order": 2},
        {"name": "allocated_to", "ownership": "system", "data_type": "string", "required": False, "output_order": 3},
        {"name": "run_id", "ownership": "system", "data_type": "string", "required": False, "output_order": 4},
    ]
    config["input_columns"] = [
        {"name": "product_id", "column": "Product ID", "ownership": "source", "data_type": "string", "required": True},
        {"name": "pt", "column": "PT", "ownership": "source", "data_type": "string", "required": False},
    ]
    config["filename"] = {"pattern": "{PROGRAM}_{ASSOCIATE_ID}_{ASSOCIATE_NAME}_{RUN_ID}.xlsx"}
    return config


class DistributionServiceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        database = Database(Path(self.tempdir.name) / "distribution.db")
        database.initialize_schema()
        programs = ProgramRepository(database)
        self.runs = RunRepository(database)
        self.snapshots = SnapshotRepository(database)
        self.audit_repository = AuditRepository(database)
        audit = AuditService(self.audit_repository, "Test Application")
        ProgramConfigurationService(programs).create_program("MX-PT", "MX PT")
        ProgramConfigurationService(programs).save_version(distribution_config())

        self.orchestration = RunOrchestrationService(runs=self.runs, snapshots=self.snapshots, manifests=ManifestRepository(database), audit=audit)
        self.file_artifacts = FileArtifactManager(base_directory=Path(self.tempdir.name) / "artifacts", artifacts=ArtifactRepository(database))
        self.source_import = SourceImportService(snapshots=self.snapshots, file_artifacts=self.file_artifacts, audit=audit)
        populations = EligiblePopulationRepository(database)
        self.eligible_population_service = EligiblePopulationService(runs=self.runs, snapshots=self.snapshots, populations=populations, audit=audit)
        self.sampling_results = SamplingResultRepository(database)
        self.sampling_service = SamplingService(runs=self.runs, snapshots=self.snapshots, populations=populations, sampling_results=self.sampling_results, audit=audit)
        self.allocation_results = AllocationResultRepository(database)
        self.allocation_service = AllocationService(runs=self.runs, snapshots=self.snapshots, sampling_results=self.sampling_results, allocation_results=self.allocation_results, audit=audit)
        self.distribution_service = DistributionService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, source_import=self.source_import, file_artifacts=self.file_artifacts, audit=audit)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_source_csv(self, row_count: int) -> Path:
        path = Path(self.tempdir.name) / "source.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, row_count + 1):
                writer.writerow([f"P{i:03d}", "Shoes"])
        return path

    def _run_to_allocated(self, *, associates: list[dict], sample_count: int, row_count: int = 20) -> str:
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=distribution_config(), sampling={"method": "count", "value": sample_count}, random_seed="seed", associates=associates)
        source_path = self._write_source_csv(row_count)
        canonical_rows, _ = self.source_import.import_source(run_id=run.run_id, file_path=source_path)
        self.eligible_population_service.validate(run_id=run.run_id, rows=canonical_rows)
        self.eligible_population_service.freeze(run_id=run.run_id, rows=canonical_rows)
        self.sampling_service.sample(run_id=run.run_id)
        self.allocation_service.finalize(run_id=run.run_id)
        return run.run_id

    def test_distribute_writes_one_workbook_per_associate_with_planned_items(self) -> None:
        associates = [
            {"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10},
            {"associate_id": "A002", "name": "John Smith", "email": "john.smith@example.test", "active": True, "target": 10, "maximum_capacity": 10},
        ]
        run_id = self._run_to_allocated(associates=associates, sample_count=20)
        artifacts = self.distribution_service.distribute(run_id=run_id)
        self.assertEqual(len(artifacts), 2)
        self.assertEqual(self.runs.get(run_id).state, RunState.DISTRIBUTED)
        filenames = {artifact.original_filename for artifact in artifacts}
        self.assertEqual(filenames, {f"MX-PT_A001_Jane_Doe_{run_id}.xlsx", f"MX-PT_A002_John_Smith_{run_id}.xlsx"})

    def test_workbook_contents_are_a_real_readable_xlsx(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5}]
        run_id = self._run_to_allocated(associates=associates, sample_count=5)
        artifacts = self.distribution_service.distribute(run_id=run_id)
        workbook = load_workbook(self.file_artifacts.run_directory(run_id) / artifacts[0].relative_path)
        data_sheet = workbook["Data"]
        header_row = next(data_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        self.assertEqual(header_row, ("Product ID", "PT", "Partner Feedback", "Allocated To", "Run ID"))
        first_data_row = next(data_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(first_data_row[3], "A001")
        self.assertEqual(first_data_row[4], run_id)
        metadata_sheet = workbook["Metadata"]
        metadata = {row[0]: row[1] for row in metadata_sheet.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(metadata["Associate ID"], "A001")
        self.assertEqual(metadata["Associate Name"], "Jane Doe")

    def test_associate_with_zero_planned_count_gets_no_file(self) -> None:
        associates = [
            {"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5},
            {"associate_id": "A002", "name": "Idle Associate", "email": "idle@example.test", "active": False, "target": 100, "maximum_capacity": 100},
        ]
        run_id = self._run_to_allocated(associates=associates, sample_count=5)
        artifacts = self.distribution_service.distribute(run_id=run_id)
        self.assertEqual(len(artifacts), 1)

    def test_requires_allocated_state(self) -> None:
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=distribution_config(), sampling={"method": "count", "value": 1}, random_seed="seed", associates=[])
        with self.assertRaises(InvalidStateTransitionError):
            self.distribution_service.distribute(run_id=run.run_id)

    def test_records_audit_event(self) -> None:
        associates = [{"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 5, "maximum_capacity": 5}]
        run_id = self._run_to_allocated(associates=associates, sample_count=5)
        self.distribution_service.distribute(run_id=run_id)
        actions = [event["action"] for event in self.audit_repository.for_run(run_id)]
        self.assertIn("RUN_DISTRIBUTED", actions)
