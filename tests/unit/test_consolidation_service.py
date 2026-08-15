from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.domain.exceptions import ConsolidationBlockedByExceptionsError, InvalidOverrideError
from operations_allocation.domain.models import RunState
from operations_allocation.infrastructure.file_artifacts import FileArtifactManager
from operations_allocation.infrastructure.xlsx_writer import write_associate_workbook
from operations_allocation.persistence.database import Database
from operations_allocation.persistence.repositories import (
    AllocationResultRepository,
    ArtifactRepository,
    AuditRepository,
    EligiblePopulationRepository,
    ManifestRepository,
    ProgramRepository,
    RunRepository,
    SamplingResultRepository,
    SnapshotRepository,
)
from operations_allocation.services.allocation import AllocationService
from operations_allocation.services.audit import AuditService
from operations_allocation.services.consolidation import ConsolidationService
from operations_allocation.services.distribution import DistributionService
from operations_allocation.services.eligible_population import EligiblePopulationService
from operations_allocation.services.program_configuration import ProgramConfigurationService
from operations_allocation.services.run_orchestration import RunOrchestrationService
from operations_allocation.services.sampling import SamplingService
from operations_allocation.services.source_import import SourceImportService
from tests.unit.test_distribution_service import distribution_config


def _read_distributed(path: Path) -> tuple[dict, tuple, list]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    metadata_sheet = workbook["Metadata"]
    metadata = {row[0]: row[1] for row in metadata_sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None}
    data_sheet = workbook["Data"]
    rows_iterator = data_sheet.iter_rows(values_only=True)
    headers = next(rows_iterator)
    rows = [list(row) for row in rows_iterator]
    workbook.close()
    return metadata, headers, rows


def _write_returned(path: Path, metadata: dict, headers: tuple, rows: list) -> None:
    content = write_associate_workbook(metadata=metadata, headers=headers, rows=rows)
    path.write_bytes(content)


class ConsolidationServiceTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        database = Database(Path(self.tempdir.name) / "consolidation.db")
        database.initialize_schema()
        programs = ProgramRepository(database)
        self.runs = RunRepository(database)
        self.snapshots = SnapshotRepository(database)
        self.audit_repository = AuditRepository(database)
        audit = AuditService(self.audit_repository, "Test Application")
        ProgramConfigurationService(programs).create_program("MX-PT", "MX PT")
        ProgramConfigurationService(programs).save_version(distribution_config())

        self.orchestration = RunOrchestrationService(runs=self.runs, snapshots=self.snapshots, manifests=ManifestRepository(database), audit=audit)
        self.file_artifacts = FileArtifactManager(base_directory=Path(self.tempdir.name) / "artifacts", artifacts=ArtifactRepository(database))
        self.source_import = SourceImportService(snapshots=self.snapshots, file_artifacts=self.file_artifacts, audit=audit)
        populations = EligiblePopulationRepository(database)
        self.eligible_population_service = EligiblePopulationService(runs=self.runs, snapshots=self.snapshots, populations=populations, audit=audit)
        self.sampling_results = SamplingResultRepository(database)
        self.sampling_service = SamplingService(runs=self.runs, snapshots=self.snapshots, populations=populations, sampling_results=self.sampling_results, audit=audit)
        self.allocation_results = AllocationResultRepository(database)
        self.allocation_service = AllocationService(runs=self.runs, snapshots=self.snapshots, sampling_results=self.sampling_results, allocation_results=self.allocation_results, audit=audit)
        self.distribution_service = DistributionService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, source_import=self.source_import, file_artifacts=self.file_artifacts, audit=audit)
        self.consolidation_service = ConsolidationService(runs=self.runs, snapshots=self.snapshots, allocation_results=self.allocation_results, file_artifacts=self.file_artifacts, audit=audit)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def _write_source_csv(self, row_count: int) -> Path:
        path = Path(self.tempdir.name) / "source.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Product ID", "PT"])
            for i in range(1, row_count + 1):
                writer.writerow([f"P{i:03d}", "Shoes"])
        return path

    def _run_to_distributed(self, *, associates: list[dict], sample_count: int, row_count: int) -> tuple[str, dict]:
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=distribution_config(), sampling={"method": "count", "value": sample_count}, random_seed="seed", associates=associates)
        source_path = self._write_source_csv(row_count)
        canonical_rows, _ = self.source_import.import_source(run_id=run.run_id, file_path=source_path)
        self.eligible_population_service.validate(run_id=run.run_id, rows=canonical_rows)
        self.eligible_population_service.freeze(run_id=run.run_id, rows=canonical_rows)
        self.sampling_service.sample(run_id=run.run_id)
        self.allocation_service.finalize(run_id=run.run_id)
        artifacts = self.distribution_service.distribute(run_id=run.run_id)
        paths_by_associate = {a.associate_id: self.file_artifacts.run_directory(run.run_id) / a.relative_path for a in artifacts}
        return run.run_id, paths_by_associate

    def _associates(self) -> list[dict]:
        return [
            {"associate_id": "A001", "name": "Jane Doe", "email": "jane.doe@example.test", "active": True, "target": 10, "maximum_capacity": 10},
            {"associate_id": "A002", "name": "John Smith", "email": "john.smith@example.test", "active": True, "target": 10, "maximum_capacity": 10},
        ]

    def test_perfect_return_has_no_open_exceptions_and_finalizes_without_override(self) -> None:
        run_id, paths = self._run_to_distributed(associates=self._associates(), sample_count=20, row_count=20)
        returned_dir = Path(self.tempdir.name) / "returned"
        returned_dir.mkdir()
        returned_paths = []
        for associate_id, path in paths.items():
            metadata, headers, rows = _read_distributed(path)
            returned_path = returned_dir / path.name
            _write_returned(returned_path, metadata, headers, rows)
            returned_paths.append((returned_path, associate_id))

        payload = self.consolidation_service.import_returned_files(run_id=run_id, files=returned_paths)
        self.assertFalse(payload["summary"]["has_open_critical_exceptions"])
        self.assertEqual(self.runs.get(run_id).state, RunState.RETURNED)

        export_artifact = self.consolidation_service.finalize(run_id=run_id)
        self.assertEqual(self.runs.get(run_id).state, RunState.CONSOLIDATED)
        workbook = load_workbook(self.file_artifacts.run_directory(run_id) / export_artifact.relative_path)
        consolidated_rows = list(workbook["Consolidated"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(consolidated_rows), 20)
        self.assertEqual(list(workbook["Quarantined"].iter_rows(min_row=2, values_only=True)), [])

    def test_missing_duplicate_unexpected_and_wrong_associate_all_detected(self) -> None:
        run_id, paths = self._run_to_distributed(associates=self._associates(), sample_count=20, row_count=30)
        returned_dir = Path(self.tempdir.name) / "returned"
        returned_dir.mkdir()

        metadata_a, headers, rows_a = _read_distributed(paths["A001"])
        wrong_associate_row = rows_a[0]  # A001's item -- will be returned by A002 instead, and NOT by A001, so it is a genuine wrong-associate case rather than colliding with a duplicate.
        rows_a = rows_a[1:]
        returned_a_path = returned_dir / paths["A001"].name
        _write_returned(returned_a_path, metadata_a, headers, rows_a)

        metadata_b, _, rows_b = _read_distributed(paths["A002"])
        rows_b = rows_b[:-1]  # Drop the last row -> 1 genuinely missing item.
        rows_b.append(list(rows_b[0]))  # Duplicate the first remaining row.
        rows_b.append(["ZZZ-UNKNOWN", "Shoes", None, "A002", run_id])  # Never in the source at all -> unexpected.
        rows_b.append(list(wrong_associate_row))  # An item truly allocated to A001, returned in B's file instead.
        returned_b_path = returned_dir / paths["A002"].name
        _write_returned(returned_b_path, metadata_b, headers, rows_b)

        payload = self.consolidation_service.import_returned_files(run_id=run_id, files=[(returned_a_path, "A001"), (returned_b_path, "A002")])
        summary = payload["summary"]
        self.assertTrue(summary["has_open_critical_exceptions"])
        self.assertEqual(len(summary["missing_identifiers"]), 1)  # The row dropped from B's file, never returned by anyone.
        self.assertGreaterEqual(summary["duplicate_count"], 1)
        self.assertGreaterEqual(summary["unexpected_count"], 1)
        self.assertGreaterEqual(summary["wrong_associate_count"], 1)
        self.assertGreaterEqual(summary["identity_issue_count"], 1)  # The wrong-associate row also fails the data-level Allocated To check.

        with self.assertRaises(ConsolidationBlockedByExceptionsError):
            self.consolidation_service.finalize(run_id=run_id)
        with self.assertRaises(InvalidOverrideError):
            self.consolidation_service.finalize(run_id=run_id, override=True)

        export_artifact = self.consolidation_service.finalize(run_id=run_id, override=True, overridden_by="qa_lead", override_reason="Investigated separately; proceeding.")
        self.assertEqual(self.runs.get(run_id).state, RunState.CONSOLIDATED)
        workbook = load_workbook(self.file_artifacts.run_directory(run_id) / export_artifact.relative_path)
        self.assertGreater(len(list(workbook["Quarantined"].iter_rows(min_row=2, values_only=True))), 0)

        actions = [event["action"] for event in self.audit_repository.for_run(run_id)]
        self.assertIn("RUN_RETURNED", actions)
        self.assertIn("RUN_CONSOLIDATED", actions)

    def test_filename_mismatch_flagged_as_identity_issue(self) -> None:
        run_id, paths = self._run_to_distributed(associates=self._associates(), sample_count=20, row_count=20)
        returned_dir = Path(self.tempdir.name) / "returned"
        returned_dir.mkdir()
        metadata, headers, rows = _read_distributed(paths["A001"])
        wrong_name_path = returned_dir / "renamed_file.xlsx"
        _write_returned(wrong_name_path, metadata, headers, rows)
        other_path = returned_dir / paths["A002"].name
        metadata_b, _, rows_b = _read_distributed(paths["A002"])
        _write_returned(other_path, metadata_b, headers, rows_b)

        payload = self.consolidation_service.import_returned_files(run_id=run_id, files=[(wrong_name_path, "A001"), (other_path, "A002")])
        self.assertGreaterEqual(payload["summary"]["identity_issue_count"], 1)

    def test_requires_distributed_state(self) -> None:
        run = self.orchestration.create_run(program_id="MX-PT", created_by="user", created_on=date(2026, 8, 15))
        self.orchestration.freeze_setup(run_id=run.run_id, program_configuration=distribution_config(), sampling={"method": "count", "value": 1}, random_seed="seed", associates=[])
        from operations_allocation.domain.exceptions import InvalidStateTransitionError

        with self.assertRaises(InvalidStateTransitionError):
            self.consolidation_service.import_returned_files(run_id=run.run_id, files=[])
