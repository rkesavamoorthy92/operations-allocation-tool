"""Reads .xlsx and .csv files into a raw, unmapped :class:`RawTable`
(ARCHITECTURE.md section 6.2). V1 supports .xlsx and .csv only; .xls is
deferred (PROJECT_SPEC.md section 5).

Cell values are converted to strings deliberately rather than left as
whatever type openpyxl/csv produced, matching the domain rule that
identifiers (and all canonical fields) are strings internally
(PROJECT_SPEC.md section 7). Integral floats are rendered without a
trailing ``.0`` and never in scientific notation. This module cannot
*recover* leading zeros Excel already stripped by auto-formatting a
column as Number -- that loss happens before this code ever sees the
cell -- but it never introduces additional loss of its own.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from operations_allocation.core.column_mapping import RawTable
from operations_allocation.domain.exceptions import UnsupportedFileFormatError

SUPPORTED_SUFFIXES = frozenset({".csv", ".xlsx"})


def read_raw_table(path: Path | str) -> RawTable:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(file_path)
    if suffix == ".xlsx":
        return _read_xlsx(file_path)
    if suffix == ".xls":
        raise UnsupportedFileFormatError("'.xls' files are not supported. Please save the file as .xlsx or .csv.")
    raise UnsupportedFileFormatError(f"Unsupported file format '{suffix}'. Only .xlsx and .csv are supported in v1.")


def _read_csv(path: Path) -> RawTable:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = tuple(next(reader))
        except StopIteration:
            return RawTable(headers=(), rows=())
        rows = tuple(
            {header: (value if value != "" else None) for header, value in zip(headers, row)}
            for row in reader
            if any(cell.strip() for cell in row)
        )
    return RawTable(headers=headers, rows=rows)


def _read_xlsx(path: Path) -> RawTable:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iterator)
        except StopIteration:
            return RawTable(headers=(), rows=())
        headers = tuple(_stringify_cell(cell) or "" for cell in header_row)
        rows = tuple(
            {header: _stringify_cell(value) for header, value in zip(headers, raw_row)}
            for raw_row in rows_iterator
            if any(cell is not None for cell in raw_row)
        )
    finally:
        workbook.close()
    return RawTable(headers=headers, rows=rows)


def _stringify_cell(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)
