"""Writes an Associate File's content (headers/rows/metadata) to real .xlsx
bytes using openpyxl. Pure serialization -- all business content is
decided by ``core.distribution`` before this module ever runs.
"""

from __future__ import annotations

from io import BytesIO
from typing import Mapping, Sequence

from openpyxl import Workbook


def write_associate_workbook(*, metadata: Mapping[str, str], headers: Sequence[str], rows: Sequence[Sequence[object]]) -> bytes:
    workbook = Workbook()

    metadata_sheet = workbook.active
    metadata_sheet.title = "Metadata"
    metadata_sheet.append(["Field", "Value"])
    for key, value in metadata.items():
        metadata_sheet.append([key, value])

    data_sheet = workbook.create_sheet("Data")
    data_sheet.append(list(headers))
    for row in rows:
        data_sheet.append(list(row))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_multi_sheet_workbook(sheets: Mapping[str, tuple[Sequence[str], Sequence[Sequence[object]]]]) -> bytes:
    """Write a workbook with one sheet per (headers, rows) entry in
    ``sheets``, in insertion order. Used by Consolidation to produce a
    Consolidated + Quarantined workbook without duplicating the openpyxl
    plumbing in write_associate_workbook."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, (headers, rows) in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(list(headers))
        for row in rows:
            sheet.append(list(row))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
